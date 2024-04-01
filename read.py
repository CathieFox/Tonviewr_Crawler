import requests
from bs4 import BeautifulSoup
import lxml

import codecs
import os
import subprocess
import time

from datetime import datetime

import pandas
import re
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from openpyxl.styles import numbers
from openpyxl.styles import Alignment



now = datetime.now()

#file  
formatted_date = now.strftime("BOLT%m%d%H")  #BOLT 

# 构造文件名
file_name_1 = f"{formatted_date}.xlsx"


# IDM的完整路径，根据你的安装位置进行调整
idm_path = 'D:\\Program File\\Download tools\\IDM\\IDMan.exe'

# 要下载的网页URL

# 定义变量
token_id = 'EQD0vdSA_NedR9uvbgN9EikRX-suesDxGeFg69XQMavfLqIw'  #BOLT 

# 使用字符串格式化构造URL
url = f'https://tonviewer.com/{token_id}?section=holders'


# 获取当前脚本运行的目录
current_directory = os.getcwd()

#print(current_directory)

file_name = '001.html'

download_path = os.path.join(current_directory, file_name)

# 调用IDM下载文件
subprocess.Popen([idm_path, '/d', url, '/p', current_directory, '/f', file_name, '/n', '/q'])

#print("下载命令已发送到IDM。")

def wait_for_download_complete(path, timeout=300):
    start_time = time.time()
    while time.time() - start_time < timeout:
        # 检查文件是否存在
        if os.path.exists(path):
            size1 = os.path.getsize(path)
            time.sleep(1)  # 等待一秒再次检查文件大小
            size2 = os.path.getsize(path)
            if size1 == size2:  # 如果文件大小在一秒钟内没有变化，假设下载完成
                return True
        else:
            time.sleep(1)  # 如果文件尚不存在，等待一秒再检查
    return False


if wait_for_download_complete(download_path):
   print("下载完成")
else:
    print("下载未在预定时间内完成")
  
  
# 指定延时的秒数
seconds = 5

#print("开始延时...")
time.sleep(seconds)
#print(f"{seconds}秒后，延时结束。")

  
  
with codecs.open('001.html', 'r', 'utf-8') as file:
    
    soup = BeautifulSoup(file, 'lxml')
  

address0 = soup.find_all("a", attrs={"class": "t1ifzxzy"})
address1 = soup.find_all("a", attrs={"class": "n1uax993"})


with open('001.txt', 'w', encoding='utf-8') as file:
    for all_addresss0 in address0:
        for all_address1 in all_addresss0:
            file.write(all_address1.get_text() + '\n')
        print(all_address1.get_text())
  
  
with open('001.txt', 'r', encoding='utf-8') as file:
    lines = file.read().splitlines()

# 将数据分组，每三个一组
data = [lines[i:i+3] for i in range(0, len(lines), 3)]

# 转换为pandas DataFrame
df = pandas.DataFrame(data)

# 将DataFrame保存到Excel文件
df.to_excel('001.xlsx', index=False, header=False)



# 使用openpyxl调整列宽
wb = load_workbook('001.xlsx')
ws = wb.active


# 设置行高，例如设置第1行的行高为40
ws.row_dimensions[1].height = 20

# 设置列宽，例如设置A列的列宽为20


ws.insert_rows(idx=1)
ws['A1'] = "Address"
ws['B1'] = "Percentage"
ws['C1'] = "Position"


for row in range(2, ws.max_row + 1):
    cell_value = str(ws[f'C{row}'].value)
    
    # 使用正则表达式移除非数字字符
    only_numbers = re.sub("[^0-9.]", "", cell_value)
    
    # 将处理后的值写回单元格
    ws[f'C{row}'].value = only_numbers
    
# 在最左边插入一列
ws.insert_cols(idx=1)

# 在新插入的列的第一行写入"Num"
ws['A1'] = "Num"

# 从第二行开始，填充1到1000
for i in range(1, 1001):
    cell_ref = f'A{i+1}'  # 因为第一行是标题，所以从A2开始填充数字
    ws[cell_ref] = i


ws.column_dimensions['A'].width = 5
ws.column_dimensions['B'].width = 50
ws.column_dimensions['C'].width = 25
ws.column_dimensions['D'].width = 25


# 定义边框的样式
thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

# 创建一个居中对齐的样式
center_aligned_text = Alignment(horizontal='center', vertical='center')


# 遍历第一列的所有单元格，并应用居中对齐
for cell in ws[1]:  # ws[1] 表示第一行
    cell.alignment = center_aligned_text


# 遍历第一到第四列的所有单元格，并应用居中对齐
for col in range(3, 5):  # 列范围从1到4
    for row in range(1, ws.max_row + 1):  # 行范围从1到最大行数
        cell = ws.cell(row=row, column=col)  # 获取当前行和列的单元格
        cell.alignment = center_aligned_text




# 遍历工作表中的所有单元格，并为每个单元格设置边框
for row in ws.iter_rows():
    for cell in row:
        cell.border = thin_border



# 保存对工作簿的更改
wb.save('001.xlsx')
os.remove('001.txt')
os.remove('001.html')


os.rename('001.xlsx',file_name_1)
